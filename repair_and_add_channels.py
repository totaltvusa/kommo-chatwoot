#!/usr/bin/env python3
import os
import json
import requests
import openpyxl
from dotenv import load_dotenv

# Load environment
load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))

KOMMO_SUBDOMAIN = os.getenv("KOMMO_SUBDOMAIN")
KOMMO_TOKEN = os.getenv("KOMMO_ACCESS_TOKEN")
CHATWOOT_BASE = os.getenv("CHATWOOT_BASE_URL").rstrip("/")
CHATWOOT_ACCT = os.getenv("CHATWOOT_ACCOUNT_ID")
CHATWOOT_TOKEN = os.getenv("CHATWOOT_API_TOKEN")

STATE_FILE = "output/migration_state.json"
EXCEL_FILE = "output/migration_log.xlsx"

# Map Kommo origins to standard slugs
ORIGIN_MAP = {
    "telegram": "channel-telegram",
    "facebook": "channel-facebook",
    "instagram_business": "channel-instagram",
    "waba": "channel-whatsapp-api",
    "com.amocrm.amocrmwa": "channel-whatsapp-lite",
}

def get_kommo_channel(lead_id: str) -> str:
    url = f"https://{KOMMO_SUBDOMAIN}.kommo.com/api/v4/talks"
    headers = {"Authorization": f"Bearer {KOMMO_TOKEN}", "Accept": "application/json"}
    params = {"filter[entity_type]": "lead", "filter[entity_id][]": lead_id}
    try:
        r = requests.get(url, headers=headers, params=params, timeout=15)
        if r.status_code == 200:
            talks = r.json().get("_embedded", {}).get("talks", [])
            if talks:
                origin = talks[0].get("origin", "")
                return ORIGIN_MAP.get(origin, f"channel-{origin}" if origin else "channel-unknown")
    except Exception as e:
        print(f"Error fetching talk for lead {lead_id}: {e}")
    return "channel-unknown"

def main():
    if not os.path.exists(STATE_FILE):
        print("State file not found.")
        return

    with open(STATE_FILE, "r") as f:
        state = json.load(f)

    # Load Excel if it exists
    wb = None
    if os.path.exists(EXCEL_FILE):
        wb = openpyxl.load_workbook(EXCEL_FILE)

    chatwoot_headers = {
        "api_access_token": CHATWOOT_TOKEN,
        "Content-Type": "application/json",
        "Accept": "application/json"
    }

    leads = state.get("funnels", {}).get("TotalTv USA", {}).get("leads", {})
    total = len(leads)
    print(f"Repairing {total} leads...")

    for idx, (lead_id, info) in enumerate(leads.items(), 1):
        if info.get("status") != "success":
            continue

        cid = info.get("chatwoot_conversation_id")
        stage_name = info.get("stage_name_corrected")
        
        # 1. Fetch channel from Kommo
        channel_label = get_kommo_channel(lead_id)
        
        # 2. Compile list of 3 labels
        funnel_label = info.get("labels", [None])[0] or "funnel-totaltv-usa"
        stage_label = info.get("labels", [None, None])[1] or "stage-unknown"
        all_labels = [funnel_label, stage_label, channel_label]
        
        print(f"[{idx}/{total}] Lead #{lead_id} -> Conv #{cid} | Channel: {channel_label}")

        # 3. Apply all 3 labels to conversation
        cw_url = f"{CHATWOOT_BASE}/api/v1/accounts/{CHATWOOT_ACCT}/conversations/{cid}/labels"
        try:
            r = requests.post(cw_url, headers=chatwoot_headers, json={"labels": all_labels}, timeout=15)
            if r.status_code in (200, 201):
                # Update state file local representation
                info["labels"] = all_labels
                info["channel_label"] = channel_label
            else:
                print(f"  [ERROR] Failed to apply labels: HTTP {r.status_code} - {r.text}")
        except Exception as e:
            print(f"  [ERROR] Exception applying labels: {e}")

        # 4. Update Excel sheet
        if wb and "TotalTv USA" in wb.sheetnames:
            ws = wb["TotalTv USA"]
            # Find row by lead_id (Col A)
            for row in range(2, ws.max_row + 1):
                if str(ws.cell(row=row, column=1).value) == str(lead_id):
                    # Column 9 is Labels Applied
                    ws.cell(row=row, column=9, value=", ".join(all_labels))
                    break

    # Save state
    with open(STATE_FILE, "w") as f:
        json.dump(state, f, indent=2)

    # Save Excel
    if wb:
        try:
            wb.save(EXCEL_FILE)
            print("Successfully updated migration_log.xlsx")
        except PermissionError:
            print("Error: Could not save migration_log.xlsx. Please make sure the file is closed.")

    print("\n[+] Repair finished. All conversations patched with Funnel, Stage, and Channel labels.")

if __name__ == "__main__":
    main()
