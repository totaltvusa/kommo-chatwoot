#!/usr/bin/env python3
"""
Kommo CRM Funnel Exporter
=========================
Exports contact and lead details from a specific Kommo funnel directly to an Excel file
without performing any migration to Chatwoot.

Usage:
  python3 export_funnel.py --funnel "Funnel Name" [--output ./output/export_log.xlsx]
"""

import os
import sys
import argparse
import datetime
import logging
import openpyxl
from dotenv import load_dotenv

from config import Config
from core.stage_resolver import StageResolver
from core.kommo_client import KommoClient

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger("export")

def print_banner(title: str):
    width = 80
    print("\n" + "=" * width)
    print(f" {title.upper()}")
    print("=" * width)

def print_section(title: str):
    print("\n" + "-" * 80)
    print(f"[*] {title}")
    print("-" * 80)

def main():
    parser = argparse.ArgumentParser(description="Kommo CRM Funnel to Excel Exporter")
    parser.add_argument("--funnel", "-f", required=True, help="Exact name of the Kommo Funnel / Pipeline to export")
    parser.add_argument("--output", "-o", default="./output/export_log.xlsx", help="Path to the output Excel workbook")
    args = parser.parse_args()

    funnel_name = args.funnel.strip()
    output_path = args.output.strip()

    # Load environment
    load_dotenv(os.path.join(os.path.dirname(__file__), ".env"))
    Config.validate_kommo()

    print_banner(f"Exporting Kommo Funnel '{funnel_name}' to Excel")

    # 1. Initialize Kommo Client & Stage Resolver
    kommo = KommoClient(
        subdomain=Config.KOMMO_SUBDOMAIN,
        access_token=Config.KOMMO_ACCESS_TOKEN
    )
    stage_resolver = StageResolver(Config.STAGE_OVERRIDES_PATH)

    # 2. Resolve Pipeline from Kommo
    print_section(f"Resolving Pipeline '{funnel_name}' from Kommo API")
    pipeline = kommo.get_pipeline_by_name(funnel_name)
    if not pipeline:
        print(f"\n[!] Error: Pipeline '{funnel_name}' not found in Kommo.")
        print("    Available pipelines:")
        for p in kommo.get_pipelines():
            print(f"      - \"{p.get('name')}\" (ID: {p.get('id')})")
        sys.exit(1)

    pipeline_id = pipeline["id"]
    statuses_map = pipeline["statuses"] # Dict[status_id, status_name]
    print(f"[+] Found Pipeline '{pipeline['name']}' (ID: {pipeline_id}) with {len(statuses_map)} stages.")

    # 3. Fetch all Leads in Pipeline
    print_section(f"Fetching Leads for Funnel '{funnel_name}'")
    leads = kommo.get_leads_in_pipeline(pipeline_id)
    total_leads = len(leads)
    print(f"[+] Found {total_leads} leads to export.")

    if total_leads == 0:
        print("\n[+] No leads found in this funnel. Export aborted.")
        sys.exit(0)

    # 4. Initialize/Load Excel Workbook
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    if os.path.exists(output_path):
        try:
            wb = openpyxl.load_workbook(output_path)
        except Exception as e:
            logger.error(f"Failed to open existing Excel file '{output_path}': {e}")
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()
        # Remove default sheet if brand new workbook
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

    # Sheet/Tab name is capped at 31 chars in Excel
    sheet_name = funnel_name[:31]
    existing_lead_ids = set()

    # Define headers
    headers = [
        "Kommo Lead ID",
        "Lead Name",
        "Contact Name",
        "Contact Phone",
        "Contact Email",
        "Funnel Name",
        "Stage (Original)",
        "Stage (Corrected)",
        "Channel Origin"
    ]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Extract lead IDs from Column A (ignoring headers)
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                try:
                    existing_lead_ids.add(int(val))
                except ValueError:
                    pass
        print(f"[*] Sheet '{sheet_name}' already exists with {len(existing_lead_ids)} leads. Resuming export...")
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)
        print(f"[+] Created new sheet '{sheet_name}' in '{output_path}'.")

    # 5. Process and Write Leads
    print_section(f"Processing and exporting {total_leads} lead(s)")
    
    for idx, lead in enumerate(leads, 1):
        lead_id = lead["id"]
        if lead_id in existing_lead_ids:
            continue
            
        lead_name = lead.get("name") or f"Lead #{lead_id}"
        status_id = lead.get("status_id")
        raw_stage_name = statuses_map.get(status_id, f"Stage #{status_id}")
        
        # Resolve stage and channel
        corrected_stage, _ = stage_resolver.resolve_stage(raw_stage_name, pipeline_name=funnel_name)
        channel = kommo.get_lead_channel(lead_id)
        
        # Fetch contact details
        contacts = lead.get("_embedded", {}).get("contacts", [])
        contact_ids = [c["id"] for c in contacts if "id" in c]
        
        contact_info = {"name": "", "phone": "", "email": ""}
        if contact_ids:
            primary_contact_id = contact_ids[0]
            try:
                contact_info = kommo.get_contact_details(primary_contact_id)
            except Exception as e:
                logger.warning(f"Failed to fetch contact #{primary_contact_id} details: {e}")

        # Format channel label into a cleaner representation for the report
        clean_channel = channel.replace("channel-", "")

        print(f"[{idx}/{total_leads}] Exporting Lead #{lead_id} | Channel: {clean_channel}")

        # Write to sheet
        row_values = [
            lead_id,
            lead_name,
            contact_info["name"],
            contact_info["phone"],
            contact_info["email"],
            funnel_name,
            raw_stage_name,
            corrected_stage,
            clean_channel
        ]
        ws.append(row_values)

        # Progressive auto-save every 50 leads to prevent data loss in case of power cut
        if idx % 50 == 0:
            try:
                wb.save(output_path)
            except PermissionError:
                logger.warning(f"Warning: Could not perform auto-save at lead #{idx} (file might be open). Continuing in memory...")

    # 6. Save Workbook
    try:
        wb.save(output_path)
        print_banner("Export Complete")
        print(f"  * Total leads exported: {total_leads}")
        print(f"  * Output Excel file:    {output_path}")
        print(f"  * Sheet/Tab Name:       '{sheet_name}'")
        print("="*80)
    except PermissionError:
        print(f"\n[!] Error: Could not save workbook to '{output_path}'.")
        print("    Please make sure the Excel file is CLOSED and not open in another program, then try again.")
        sys.exit(1)

if __name__ == "__main__":
    main()
