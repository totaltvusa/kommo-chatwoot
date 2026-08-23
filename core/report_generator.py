import os
import sys
import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("migration.report_generator")

COLUMNS = [
    "Kommo Lead ID",
    "Contact Name",
    "Contact Phone",
    "Contact Email",
    "Funnel Name",
    "Stage",
    "Chatwoot Contact ID",
    "Chatwoot Conversation ID",
    "Labels Applied",
    "Messages Migrated",
    "Last Message Date",
    "Migration Date",
    "Status",
    "Error Detail",
]

class ReportGenerator:
    """
    Manages the migration log Excel workbook using openpyxl.
    - Maintains one sheet per funnel.
    - Updates rows in-place by Kommo Lead ID (preventing duplicates on re-runs).
    - Safely handles file locks / open spreadsheets.
    """
    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb: Optional[openpyxl.Workbook] = None
        self._ensure_workbook_loaded()

    def _ensure_workbook_loaded(self):
        self.file_path.parent.mkdir(parents=True, exist_ok=True)
        if self.file_path.exists():
            try:
                self.wb = openpyxl.load_workbook(self.file_path)
                logger.info(f"Loaded existing Excel workbook from '{self.file_path}'.")
            except Exception as e:
                logger.error(f"Failed to load existing workbook at '{self.file_path}': {e}. Creating new.")
                self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.Workbook()
            # Remove default empty sheet if creating fresh
            if "Sheet" in self.wb.sheetnames:
                self.wb.remove(self.wb["Sheet"])

    def _get_or_create_sheet(self, funnel_name: str) -> openpyxl.worksheet.worksheet.Worksheet:
        # Excel sheet titles max length is 31 characters
        sheet_title = funnel_name[:31]
        
        if sheet_title in self.wb.sheetnames:
            ws = self.wb[sheet_title]
        else:
            ws = self.wb.create_sheet(title=sheet_title)
            self._apply_header_styles(ws)
            
        return ws

    def _apply_header_styles(self, ws: openpyxl.worksheet.worksheet.Worksheet):
        """Creates and styles the header row for a sheet."""
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        ws.append(COLUMNS)
        ws.row_dimensions[1].height = 28

        for col_idx in range(1, len(COLUMNS) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Set initial column widths
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 5, 14)

    def log_lead_migration(
        self,
        funnel_name: str,
        lead_id: int,
        contact_name: str,
        contact_phone: str,
        contact_email: str,
        stage_name: str,
        chatwoot_contact_id: Optional[int],
        chatwoot_conversation_id: Optional[int],
        labels_applied: List[str],
        messages_migrated_count: int,
        last_message_date: str,
        migration_date: str,
        status: str,
        error_detail: str = ""
    ):
        """
        Updates an existing row if lead_id is present in the funnel tab,
        or appends a new row.
        """
        ws = self._get_or_create_sheet(funnel_name)
        
        row_data = [
            int(lead_id),
            str(contact_name or ""),
            str(contact_phone or ""),
            str(contact_email or ""),
            str(funnel_name),
            str(stage_name),
            chatwoot_contact_id if chatwoot_contact_id is not None else "",
            chatwoot_conversation_id if chatwoot_conversation_id is not None else "",
            ", ".join(labels_applied) if labels_applied else "",
            int(messages_migrated_count or 0),
            str(last_message_date or ""),
            str(migration_date or ""),
            str(status),
            str(error_detail or "")
        ]

        # Find existing row by Kommo Lead ID in column 1 (ignoring header)
        target_row_idx = None
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val is not None and str(cell_val).strip() == str(lead_id).strip():
                target_row_idx = row
                break

        if target_row_idx:
            # Update row in place
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=target_row_idx, column=col_idx, value=val)
            self._style_data_row(ws, target_row_idx, status)
            logger.debug(f"Updated existing row {target_row_idx} for Lead ID #{lead_id} in tab '{funnel_name}'.")
        else:
            # Append new row
            ws.append(row_data)
            new_row_idx = ws.max_row
            self._style_data_row(ws, new_row_idx, status)
            logger.debug(f"Appended new row {new_row_idx} for Lead ID #{lead_id} in tab '{funnel_name}'.")

        self.save()

    def _style_data_row(self, ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int, status: str):
        data_font = Font(name="Calibri", size=10)
        data_align = Alignment(vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='E0E0E0'),
            right=Side(style='thin', color='E0E0E0'),
            top=Side(style='thin', color='E0E0E0'),
            bottom=Side(style='thin', color='E0E0E0')
        )
        
        # Color code status column (column 13)
        status_cell = ws.cell(row=row_idx, column=13)
        if status.lower() == "success":
            status_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="385723")
        else:
            status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # soft red
            status_cell.font = Font(name="Calibri", size=10, bold=True, color="C00000")

        for col in range(1, len(COLUMNS) + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = thin_border
            if col != 13:
                c.font = data_font
            c.alignment = data_align

    def save(self):
        """Saves workbook with retry and friendly prompt if locked by Excel."""
        while True:
            try:
                self.wb.save(self.file_path)
                break
            except PermissionError:
                msg = (
                    f"\n[CRITICAL ERROR] Cannot save Excel file at '{self.file_path}'.\n"
                    f"The workbook is currently open in Excel or LibreOffice.\n"
                    f"--> Please CLOSE the file in Excel and press ENTER to retry..."
                )
                print(msg)
                input()
            except Exception as e:
                logger.error(f"Error saving workbook '{self.file_path}': {e}")
                break
